import urllib.error
import pandas as pd
from Bio import SeqIO, Entrez, Restriction
from Bio.Seq import Seq
from openpyxl import workbook, load_workbook
from pydna.dseq import Dseq
from pydna.dseqrecord import Dseqrecord
from pydna.design import primer_design
import random
from tqdm import tqdm

# Parameters

codon_list = ['ATT', 'ATC', 'ATA', 'CTT', 'CTC', 'CTA', 'CTG', 'TTA', 'TTG', 'GTT', 'GTC', 'GTA', 'GTG', 'TTT',
              'TTC', 'ATG', 'TGT', 'TGC', 'GCT', 'GCC', 'GCA', 'GCG', 'GGT', 'GGC', 'GGA', 'GGG', 'CCT', 'CCC',
              'CCA', 'CCG', 'ACT', 'ACC', 'ACA', 'ACG', 'TCT', 'TCC', 'TCA', 'TCG', 'AGT', 'AGC', 'TAT', 'TAC',
              'TGG', 'CAA', 'CAG', 'AAT', 'AAC', 'CAT', 'CAC', 'GAA', 'GAG', 'GAT', 'GAC', 'AAA', 'AAG', 'CGT',
              'CGC', 'CGA', 'CGG', 'AGA', 'AGG', 'TAA', 'TAG', 'TGA']

ProteinsSequenced = True  # False activates Entrez sequence fetching
SeqRevTranscribed = False  # False activates reverse transcription
filename = "Book2.xlsx"
Entrez.email = "royerz.2002@gmail.com"
verbose = False

digestionName1 = input('What is your first digestion site?')
digestionName2 = input('What is your second digestion site?')

digestionSite1 = getattr(Restriction, digestionName1)
digestionSite2 = getattr(Restriction, digestionName2)


class DNA:
    def __init__(self, seq, dseq):
        self.seq = seq
        self.dseq = dseq


class Plasmid(DNA):
    def __init__(self, restriction_sites, linearized):
        self.restriction_sites = restriction_sites
        self.linearized = linearized


class Insert(DNA):
    def __init__(self, fp, rp, id, mrna, aa):
        self.fp = fp
        self.rp = rp
        self.id = id
        self.mrna = mrna
        self.aa = aa


def aa_to_codon(n):
    switch = {
        "A": (random.choices(["GCU", "GCC", "GCA", "GCG"], weights=(0.19, 0.25, 0.22, 0.34), k=1)),
        "I": (random.choices(["AUU", "AUC", "AUA"], weights=(0.47, 0.46, 0.07), k=1)),
        "L": (random.choices(["UUA", "UUG", "CUU", "CUC", "CUA", "CUG"],
                             weights=(0.11, 0.11, 0.10, 0.10, 0.03, 0.55), k=1)),
        "M": ["AUG"],
        "V": (random.choices(["GUU", "GUC", "GUA", "GUG"], weights=(0.29, 0.20, 0.17, 0.34), k=1)),
        "F": (random.choices(["UUU", "UUC"], weights=(0.51, 0.49), k=1)),
        "W": ["UUG"],
        "Y": (random.choices(["UAU", "UAC"], weights=(0.53, 0.47), k=1)),
        "N": (random.choices(["AAU", "AAC"], weights=(0.39, 0.61), k=1)),
        "C": (random.choices(["UGU", "UGC"], weights=(0.43, 0.57), k=1)),
        "Q": (random.choices(["CAA", "CAG"], weights=(0.31, 0.69), k=1)),
        "S": (random.choices(["UCU", "UCC", "UCA", "UCG"], weights=(0.19, 0.17, 0.12, 0.13), k=1)),
        "T": (random.choices(["ACU", "ACC", "ACA", "ACG"], weights=(0.16, 0.47, 0.13, 0.24), k=1)),
        "D": (random.choices(["GAU", "GAC"], weights=(0.59, 0.41), k=1)),
        "E": (random.choices(["GAA", "GAG"], weights=(0.7, 0.3), k=1)),
        "R": (random.choices(["CCU", "CGC", "CGA", "CGG"], weights=(0.42, 0.37, 0.05, 0.08), k=1)),
        "H": (random.choices(["AAU", "AAC"], weights=(0.39, 0.61), k=1)),
        "K": (random.choices(["AAA", "AAG"], weights=(0.76, 0.24), k=1)),
        "G": (random.choices(["GGU", "GGC", "GGA", "GGG"], weights=(0.38, 0.40, 0.09, 0.13), k=1)),
        "P": (random.choices(["CCU", "CCC", "CCA", "CCG"], weights=(0.16, 0.10, 0.2, 0.54), k=1)),
    }

    codon_elem = switch[n]
    codon_str = f"{codon_elem[0]}"

    return codon_str


def id_to_aa_sequence(protein_id):
    try:
        # Get AA sequence from Entrez database.
        handle = Entrez.efetch(db="protein", id=protein_id, rettype="fasta")
        record = SeqIO.read(handle, "fasta")

        aa_sequence = record.seq
        if verbose: print(f"AA Sequence: {aa_sequence}")
        ws[f'D{i + 2}'] = f"{aa_sequence}"

        return aa_sequence

    except urllib.error.HTTPError:  # If cannot find AA sequence, fill all columns in excel "n/a".
        if verbose: print(f"HTTP Error at position {i}.")
        ws[f'C{i + 2}'] = "n/a"
        ws[f'D{i + 2}'] = "n/a"
        ws[f'E{i + 2}'] = "n/a"
        ws[f'F{i + 2}'] = "n/a"
        ws[f'G{i + 2}'] = "n/a"
        ws[f'H{i + 2}'] = "n/a"

    except Exception as e:  # Handle unexpected EOF and continue loop.
        print(e)


def aa_to_mrna(aa_sequence):
    try:
        # Reverse translate Protein to RNA
        codon_list = list(map(aa_to_codon, aa_sequence))  # Use RT to get codon list from AAs)
        if verbose: print(f"Codons: {codon_list}")

        mrna = ''.join(codon_list)  # Merge codons to get mRNA strand.
        if verbose: print("RNA: " + mrna)

        return mrna

    except Exception as e:
        print(e)


def mrna_to_cdna_seq(mrna):
    try:
        cdna = mrna.replace("U", "T")  # Reverse translate mRNA to Crick (anti-coding) strand.
        cdna_seq = Seq(cdna)  # Convert cDNA sequence to Seq data structure.
        if verbose: print("cDNA (5-3): " + cdna_seq)
        if verbose: print("cDNA (3-5): " + cdna_seq.complement())
        ws[f'E{i + 2}'] = f"{cdna_seq.complement()}"  # Send cDNA Watson (coding) strand to excel sheet.
        ws[f'F{i + 2}'] = f"{cdna_seq}"  # Send cDNA Crick (anti-coding) strand to excel sheet.
        return cdna_seq

    except Exception as e:
        print(e)


def cdna_to_primer(cdna_seq):
    try:
        # Design primers.
        cdna_dseq = Dseqrecord(cdna_seq, str(cdna_seq.complement), linear=True, circular=False)
        ampl = primer_design(cdna_dseq)

        fw = ampl.forward_primer
        rev = ampl.reverse_primer[::-1]

        ws[f'G{i + 2}'] = f"{str(fw.seq)}"  # Send forward primer to excel sheet.
        ws[f'H{i + 2}'] = f"{str(rev.seq)}"  # Send reverse primer to excel sheet.

        primers = (fw, rev)

        return primers

    except Exception as e:
        print(e)


def restriction_assembly(insert, restriction1, restriction2):
    fw, rev = cdna_to_primer(insert)

    # Check for hairpin structure in primer.

    for codon in codon_list:
        codon = codon
        anti_codon = str(Seq(codon).complement())
        if fw.find(codon) == -1 and rev.find(anti_codon) == -1:  # check if codon in ForwardPrimer and ReversePrimer
            break

    fwRestricting = codon + fw + getattr(Restriction, restriction1)
    revRestricting = anti_codon + rev + getattr(Restriction, restriction2)

    ws[f'I{i + 2}'] = f"{str(fwRestricting.seq)}"  # Send forward restricting primer to excel sheet.
    ws[f'J{i + 2}'] = f"{str(revRestricting.seq)}"  # Send reverse restricting primer to excel sheet.

    return fwRestricting, revRestricting


wb = load_workbook(filename)
sheets = wb.sheetnames
ws = wb[sheets[0]]

dataframe = pd.DataFrame(ws.values, columns=["prot", "spp", "code", "aa", "cDNA(5-3)",
                                             "cDNA(3-5)", "primer", "rev_primer"])
dataframe.drop(index=dataframe.index[0], axis=0, inplace=True)

print("Starting protein fetching, reverse translation and primer design.")

code_list = dataframe.code.tolist()

if verbose:
    print(code_list)
    print(len(code_list))

for i in tqdm(range(len(code_list))):
    aa_sequence = id_to_aa_sequence(code_list[i])
    mrna = aa_to_mrna(aa_sequence)
    cdna_seq = mrna_to_cdna_seq(mrna)
    primers = cdna_to_primer(cdna_seq)

print("Protein fetching complete, saving excel sheet.")
wb.save(filename=filename)
